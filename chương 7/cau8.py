from openpyxl import load_workbook
import os
import sys

FILE_PATH = "demo.xlsx"

def doc_du_lieu_excel(filename):
   
    if not os.path.exists(filename):
        print(f"Lỗi: Không tìm thấy file '{filename}'. Vui lòng chạy lại chương trình Câu 7 để tạo file.")
        return

    print(f"--- BẮT ĐẦU ĐỌC DỮ LIỆU TỪ {filename} ---")
    
    try:
        # 1. Tải workbook (wb)
        wb = load_workbook(filename)
        
        # 2. Lấy tên sheet đầu tiên và chọn worksheet (ws)
        ws_name = wb.sheetnames[0]
        ws = wb[ws_name]
        
        print(f"Đọc từ Sheet: {ws.title}")

        # 3. Duyệt qua tất cả các dòng dữ liệu (sử dụng ws.values là cách hiệu quả nhất)
        for row in ws.values:
            line_output = ""
            for value in row:
                # Thêm tab (\t) để phân cách cột
                line_output += f"{value}\t" 
                
            # In ra toàn bộ dòng
            print(line_output)
            
        print("\n--- HOÀN TẤT ĐỌC FILE ---")

    except Exception as e:
        print(f"Lỗi: Không thể xử lý file Excel. Lỗi: {e}")
        
if __name__ == "__main__":
    doc_du_lieu_excel(FILE_PATH)