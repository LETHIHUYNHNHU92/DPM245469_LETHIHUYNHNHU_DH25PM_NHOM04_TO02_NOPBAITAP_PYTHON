import openpyxl
from openpyxl import Workbook, load_workbook
import os
import operator

FILE_PATH = "quanly_nhanvien.xlsx"
HEADER = ['STT', 'Mã', 'Tên', 'Tuổi']

def KhoiTaoFile(path):
    """Khởi tạo file Excel với header nếu file chưa tồn tại."""
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = "NhanVienData"
        # Ghi header vào dòng đầu tiên
        ws.append(HEADER)
        wb.save(path)
        print(f"✅ Đã tạo file '{path}' với tiêu đề.")

def NhapNhanVienVaLuu(path):
    """Cho phép người dùng nhập thông tin và lưu vào file Excel (append)."""
    
    # 1. Nhập dữ liệu từ bàn phím
    print("\n--- NHẬP DỮ LIỆU NHÂN VIÊN ---")
    ma_nv = input("Nhập Mã NV: ").strip()
    ten_nv = input("Nhập Tên NV: ").strip()
    
    while True:
        try:
            tuoi = int(input("Nhập Tuổi: "))
            break
        except ValueError:
            print("Tuổi phải là một số nguyên hợp lệ.")
            
    # 2. Đọc file để xác định STT và lưu
    try:
        wb = load_workbook(path)
        ws = wb.active
        
        # STT mới = Số dòng hiện tại (cộng thêm 1) - 1 (trừ đi dòng header)
        stt_moi = ws.max_row
        
        # Dữ liệu mới: [STT, Mã, Tên, Tuổi]
        new_row = [stt_moi, ma_nv, ten_nv, tuoi]
        
        # Ghi vào dòng cuối cùng
        ws.append(new_row)
        wb.save(path)
        print(f"Đã lưu nhân viên {ten_nv} thành công.")
        
    except Exception as e:
        print(f"Lỗi khi lưu vào file: {e}")

def DocVaSapXep(path, ascending=True):
    """Đọc dữ liệu từ file, sắp xếp theo Tuổi và xuất ra màn hình."""
    
    if not os.path.exists(path):
        print("Lỗi: File dữ liệu chưa tồn tại.")
        return

    try:
        wb = load_workbook(path)
        ws = wb.active
        
        ds_nhan_vien = []
        
        # 1. Đọc dữ liệu (Bỏ qua dòng Header đầu tiên)
        # Bắt đầu từ dòng thứ 2 (row[1:] trong ws.rows)
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            if any(row): # Đảm bảo dòng không rỗng
                # Cột Tuổi (index 3) phải là số để sắp xếp
                ds_nhan_vien.append({
                    'STT': row[0],
                    'Mã': row[1],
                    'Tên': row[2],
                    'Tuổi': int(row[3]) # Chuyển Tuổi sang int
                })
        
        # 2. Sắp xếp theo Tuổi tăng dần (Tuổi là key 'Tuổi')
        ds_nhan_vien.sort(key=operator.itemgetter('Tuổi'), reverse=not ascending)
        
        # 3. Xuất kết quả
        print("\n--- DANH SÁCH NHÂN VIÊN (ĐÃ SẮP XẾP THEO TUỔI) ---")
        print(f"{'STT':<4} | {'Mã':<6} | {'Tên':<15} | {'Tuổi':<4}")
        print("-" * 35)
        for nv in ds_nhan_vien:
            print(f"{nv['STT']:<4} | {nv['Mã']:<6} | {nv['Tên']:<15} | {nv['Tuổi']:<4}")

    except Exception as e:
        print(f"Lỗi khi đọc hoặc sắp xếp dữ liệu: {e}")
        
def main_menu():
    
    KhoiTaoFile(FILE_PATH)
    
    while True:
        print("\n" + "="*40)
        print("CÂU 11: QUẢN LÝ NHÂN VIÊN (EXCEL)")
        print("1. Nhập và Lưu Nhân Viên mới")
        print("2. Đọc & Sắp xếp Nhân Viên (Theo Tuổi tăng dần)")
        print("0. Thoát")
        print("="*40)
        
        lua_chon = input("Nhập lựa chọn của bạn: ").strip()
        
        if lua_chon == '1':
            NhapNhanVienVaLuu(FILE_PATH)
        elif lua_chon == '2':
            DocVaSapXep(FILE_PATH, ascending=True)
        elif lua_chon == '0':
            print("Đã thoát chương trình.")
            break
        else:
            print("Lựa chọn không hợp lệ. Vui lòng nhập lại.")

if __name__ == "__main__":
    main_menu()